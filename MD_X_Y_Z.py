import openpyxl
excel_file_name='work_111.xlsx'
File_In=openpyxl.load_workbook(filename=excel_file_name,read_only=False)
print('Имя конфигурационного файла?')
input_data_file_name=input()
A=[]
fin=open(input_data_file_name,'r')
for line in fin:
    A.append(line)
fin.close()

for el in A:
    Sheet_In=File_In[el.split()[0]] #имя скважины
    nrows=int(el.split()[1]) #число строк
    data_dict={}
    for i in range(2,nrows+1):
        data_dict[i]=(Sheet_In.cell(row=i,column=1).value,Sheet_In.cell(row=i,column=2).value,Sheet_In.cell(row=i,column=3).value,Sheet_In.cell(row=i,column=4).value)
    full_row=el.split()
    MDs=[]
    for i in range(0,len(full_row)):
        if i>1:
            MDs.append(float(full_row[i]))
    Zs=[]
    Xs=[]
    Ys=[]
    for i in range(0,len(MDs)):
        for key in data_dict:
            if data_dict[key][3]>=MDs[i]:
                Z=data_dict[key-1][2]+(MDs[i]-data_dict[key-1][3])*(data_dict[key][2]-data_dict[key-1][2])/(data_dict[key][3]-data_dict[key-1][3])
                Zs.append(Z)
                X=data_dict[key-1][0]+(MDs[i]-data_dict[key-1][3])*(data_dict[key][0]-data_dict[key-1][0])/(data_dict[key][3]-data_dict[key-1][3])
                Xs.append(X)
                Y=data_dict[key-1][1]+(MDs[i]-data_dict[key-1][3])*(data_dict[key][1]-data_dict[key-1][1])/(data_dict[key][3]-data_dict[key-1][3])
                Ys.append(Y)
                break
    nrow=3
    for i in range(0,len(Xs)):
        Sheet_In.cell(row=nrow,column=16).value=Xs[i]
        Sheet_In.cell(row=nrow,column=17).value=Ys[i]
        Sheet_In.cell(row=nrow,column=18).value=Zs[i]
        nrow+=1
    File_In.save(excel_file_name)
    File_In.close()
    

'''print('Имя скважины?')
wellname=input()
Sheet_In=File_In[wellname]
print('Сколько записей-строк?')
nrows=int(input())
data_dict={}
for i in range(2,nrows+1):
    data_dict[i]=(Sheet_In.cell(row=i,column=1).value,Sheet_In.cell(row=i,column=2).value,Sheet_In.cell(row=i,column=3).value,Sheet_In.cell(row=i,column=4).value)



print('Через пробел введите нужные к интерполяции MDшки')
MDs=input().split()
MDs_int=[]
for i in range(0,len(MDs)):
    MDs_int.append(float(MDs[i]))

Zs=[]
Xs=[]
Ys=[]
for i in range(0,len(MDs_int)):
    for key in data_dict:
        if data_dict[key][3]>=MDs_int[i]:
            Z=data_dict[key-1][2]+(MDs_int[i])*(data_dict[key][2]-data_dict[key-1][2])/(data_dict[key][3]-data_dict[key-1][3])
            Zs.append(Z)
            X=data_dict[key-1][0]+(MDs_int[i])*(data_dict[key][0]-data_dict[key-1][0])/(data_dict[key][3]-data_dict[key-1][3])
            Xs.append(X)
            Y=data_dict[key-1][1]+(MDs_int[i])*(data_dict[key][1]-data_dict[key-1][1])/(data_dict[key][3]-data_dict[key-1][3])
            Ys.append(Y)
            break
nrow=3
for i in range(0,len(Xs)):
    Sheet_In.cell(row=nrow,column=16).value=Xs[i]
    Sheet_In.cell(row=nrow,column=17).value=Ys[i]
    Sheet_In.cell(row=nrow,column=18).value=Zs[i]
    nrow+=1

File_In.save(excel_file_name)
File_In.close()'''