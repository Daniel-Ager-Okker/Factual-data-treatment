import openpyxl
excel_file_name='factual_incl.xlsx'
File_In=openpyxl.load_workbook(filename=excel_file_name,read_only=True)
sheets=File_In.sheetnames #список имен листов
for i in range(0,len(sheets)):
    worksheet=File_In[sheets[i]] #открываем i-тый лист
    fout=open(f"{sheets[i]}.dev",'w')
    for j in range(1,10000):
        if worksheet.cell(row=j,column=1).value is None:
            break
        else:
            fout.write(str(worksheet.cell(row=j,column=1).value)+' '+str(worksheet.cell(row=j,column=3).value)+' '+str(worksheet.cell(row=j,column=4).value)+'\n')
    fout.close()
File_In.close()
